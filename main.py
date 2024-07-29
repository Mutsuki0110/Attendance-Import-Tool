import streamlit as st
import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime, timedelta

# xlookup function
def xlookup(lookup_value, lookup_array, return_array, if_not_found:str = ''):
	match_value = return_array.loc[lookup_array == lookup_value]

	if match_value.empty:
		return '' if if_not_found == '' else if_not_found
	else:
		return match_value.tolist()[0]

# Set the page configuration
st.set_page_config(
	page_title="Attendance Import Tool",
	page_icon="🕔",
	layout="centered"
)
# Define the disclaimer warning box
st.warning("""
	### Disclaimer  
	- This tool is designed to automate the process of importing attendance records from multiple sources.  
		
	- This tool is intended for **internal use only** and should not be shared with unauthorized individuals.
	By using this tool, you agree to comply with the company's data security and privacy policies.  
		
	- Please make sure to review the imported records for accuracy before finalizing the import process.  
		
	- This tool is not affiliated with **JD Logistics United States Company** and is not under the regulation of **JD Logistics United States Company**.
	By using this tool, you agree that any unwanted results or consequences are not the responsibility of the author of this tool.  
	  
	**Using this tool indicates your acceptance of the above disclaimer.**
""")

# Set the page title
st.title("Attendance Import Tool")

# Prompt the user for the initial date
date = st.date_input("Pick a date", value=None)
# Info message for the date selection
st.info("Select the date for which you want to import attendance records.")
# LMS attendance file upload
uploaded_lms_files = st.file_uploader("Upload LMS Record Files", type=["xlsx"], accept_multiple_files=True)
# Info message for the LMS attendance files
st.info("""
		File name should have a format of `lms-record-{mm}-{dd}.xlsx`.  
		For instance, `lms-record-07-18.xlsx` is a valid filename; `lms_record_07_18.xlsx` is not.
""")
# Define the directory to save the uploaded files
source_files_folder = "source_files"

# Create the directory if it doesn't exist
os.makedirs(source_files_folder, exist_ok=True)

# Check if LMS files are uploaded
if uploaded_lms_files:
	for uploaded_lms_file in uploaded_lms_files:
		lms_file_name = uploaded_lms_file.name
		lms_file_path = os.path.join(source_files_folder, lms_file_name)
		
		# Save the uploaded file to the specified directory
		with open(lms_file_path, "wb") as f:
			f.write(uploaded_lms_file.getbuffer())
		
		st.success(f"File {lms_file_name} uploaded successfully.")
# Manual attendance file upload
uploaded_manual_attendance_file = st.file_uploader("Upload Manual Attendance File", type=["xlsx"])
# Info message for the manual attendance file
st.info("""
		File name should have a format of `manual-attendance-{mm}-{dd}.xlsx`.  
		For instance, `manual-attendance-07-18.xlsx` is a valid filename; `manual_attendance_07_18.xlsx` is not.
""")

# Check if manual attendance file is uploaded
if uploaded_manual_attendance_file:
	manual_attendance_file = uploaded_manual_attendance_file.name
	manual_attendance_path = os.path.join(source_files_folder, manual_attendance_file)
	
	# Save the uploaded file to the specified directory
	with open(manual_attendance_path, "wb") as f:
		f.write(uploaded_manual_attendance_file.getbuffer())
	
	st.success(f"File {manual_attendance_file} uploaded successfully.")
# Import template file upload
uploaded_template = st.file_uploader("Upload Import Template File", type=["xlsx"])
# Info message for the import template file
st.info("Import template retreived from LMS.")

# Define the directory to save the final output files
final_output_files_folder = "final_output_files"

# Create the directory if it doesn't exist
os.makedirs(final_output_files_folder, exist_ok=True)

# Check if import template file is uploaded
if uploaded_template:
	template_file = uploaded_template.name
	template_path = os.path.join(final_output_files_folder, template_file)
	
	# Save the uploaded file to the specified directory
	with open(template_path, "wb") as f:
		f.write(uploaded_template.getbuffer())
	
	st.success(f"File {template_file} uploaded successfully.")

# Check if the date is valid
if st.button("Process Files"):
	# Convert the date to a string with format 'mm-dd'
	date = date.strftime("%m-%d")
	# Define the initial file name
	initial_file_name = f"lms-record-{date}.xlsx"
	# Extract the date from the file name
	base_name = "lms-record-"
	date_str = date
	initial_date = datetime.strptime(date_str, "%m-%d")
	# Define the folder paths
	extracted_files_folder = "extracted_files"
	# Create the extracted_files folder if it doesn't exist
	os.makedirs(extracted_files_folder, exist_ok=True)

	# Loop through all files in the source_files folder that start with "lms-record-"
	for file_name in os.listdir(source_files_folder):
		if file_name.startswith("lms-record-") and file_name.endswith(".xlsx"):
			lms_file_path = os.path.join(source_files_folder, file_name)
			# Load the Excel file into a DataFrame
			df = pd.read_excel(lms_file_path)
			# Extract the specified columns
			extracted_df = df[['employeeNo', 'firstCheckIn', 'lastCheckOut']]
			
			# Save the extracted data to a new Excel file
			extracted_file_name = f"extracted-{file_name}"
			extracted_file_path = os.path.join(extracted_files_folder, extracted_file_name)
			extracted_df.to_excel(extracted_file_path, index=False)
			
			print(f"Processed file: {file_name}, saved extracted data to: {extracted_file_name}")

	# Define the folder paths
	combined_records_folder = "combined_records"
	manual_attendance_file = f"source_files/manual-attendance-{date}.xlsx"

	# Create the combined_records folder if it doesn't exist
	os.makedirs(combined_records_folder, exist_ok=True)

	# Initialize the Excel writer
	merged_file_path = os.path.join(combined_records_folder, "merged-manual-attendance.xlsx")
	with pd.ExcelWriter(merged_file_path) as writer:
		# Read the original manual attendance file and write it to the merged file
		manual_attendance_df = pd.read_excel(manual_attendance_file)
		manual_attendance_df.to_excel(writer, sheet_name='main', index=False)
		# Read all extracted files and write them to separate sheets
		for file_name in os.listdir(extracted_files_folder):
			if file_name.endswith(".xlsx"):
				file_path = os.path.join(extracted_files_folder, file_name)
				df = pd.read_excel(file_path)
				# Extract the date from the file name
				date_str = file_name.replace("extracted-lms-record-", "").replace(".xlsx", "")
				# Write the DataFrame to a sheet named after the date
				df.to_excel(writer, sheet_name=date_str, index=False)

	# Load the workbook to apply date formatting
	wb = load_workbook(merged_file_path)
	ws = wb['main']

	# Define the date format style
	date_style = NamedStyle(name="short_date", number_format="MM-DD")

	# Apply the date format to the first row
	for cell in ws[1]:
		if isinstance(cell.value, datetime):
			cell.style = date_style

	# Save the workbook with the updated styles
	wb.save(merged_file_path)

	print(f"Merged file saved to: {merged_file_path}")

	# Define the folder paths
	combined_records_folder = "combined_records"
	merged_file_path = os.path.join(combined_records_folder, "merged-manual-attendance.xlsx")

	# Load the workbook
	wb = load_workbook(merged_file_path)
	main_df = pd.read_excel(merged_file_path, sheet_name='main')

	# Iterate through all the daily record worksheets
	for sheet_name in wb.sheetnames:
		if sheet_name != 'main':
			daily_df = pd.read_excel(merged_file_path, sheet_name=sheet_name)
			
			# Perform the XLOOKUP-like operation and add the new column to the main DataFrame
			main_df[f'lms-start-{sheet_name}'] = main_df['employeeNo'].apply(xlookup, args=(daily_df['employeeNo'], daily_df['firstCheckIn']))
			main_df[f'lms-end-{sheet_name}'] = main_df['employeeNo'].apply(xlookup, args=(daily_df['employeeNo'], daily_df['lastCheckOut']))

	xlookup_file_path = os.path.join(combined_records_folder, "xlookup-manual-attendance.xlsx")

	# Save the updated main DataFrame back to the merged file
	with pd.ExcelWriter(xlookup_file_path) as writer:
		main_df.to_excel(writer, sheet_name='main', index=False)
		for sheet_name in wb.sheetnames:
			if sheet_name != 'main':
				daily_df = pd.read_excel(merged_file_path, sheet_name=sheet_name, engine="openpyxl")
				daily_df.to_excel(writer, sheet_name=sheet_name, index=False)

	print(f"XLOOKUP-ed file saved to: {xlookup_file_path}")

	# Define the import template file path
	import_template_file = 'final_output_files/importTemplate.xlsx'

	# Define the destination directory
	import_dest_dir = 'final_output_files'

	# Create the final_output_files folder if it doesn't exist
	os.makedirs(import_dest_dir, exist_ok=True)

	# Construct the new filename
	new_import_filename = f'importTemplate-{date}.xlsx'

	# Construct the full destination path
	import_dest_path = os.path.join(import_dest_dir, new_import_filename)

	# Copy the file to the new destination with the new name
	shutil.copy(import_template_file, import_dest_path)

	print(f'File copied to {import_dest_path}')

	# File paths
	xlookup_file_path = 'combined_records/xlookup-manual-attendance.xlsx'
	output_dir = 'final_output_files'

	# Load the xlookup manual attendance workbook
	wb = load_workbook(xlookup_file_path)
	main_df = pd.read_excel(xlookup_file_path, sheet_name='main')

	# Create a list to hold the extracted data
	extracted_data = []

	# Iterate through all the daily record worksheets
	for sheet_name in wb.sheetnames:
		if sheet_name != 'main':
			start_col = f'start-{sheet_name}'
			end_col = f'end-{sheet_name}'
			lms_start_col = f'lms-start-{sheet_name}'
			lms_end_col = f'lms-end-{sheet_name}'
			# Iterate through the main DataFrame
			for _, row in main_df.iterrows():
				# Skip rows that already have LMS start and end times
				if pd.notna(row[lms_start_col]) or pd.notna(row[lms_end_col]):
					continue
				# Check if the row has an employee number and both start and end times, and has manual recorded start and end times
				if (pd.notna(row['employeeNo'])) and (pd.isna(row[lms_start_col]) and pd.isna(row[lms_end_col])) and (pd.notna(row[start_col]) and pd.notna(row[end_col])):
					employee_no = row['employeeNo']
					# get the start and end times from manually recorded columns
					start_time = row[start_col] if pd.isna(row[lms_start_col]) else ''
					end_time = row[end_col] if pd.isna(row[lms_end_col]) else ''
					# Check if both start_time and end_time are not empty
					if start_time and end_time:
						start_time_dt = datetime.strptime(start_time, "%H:%M:%S")
						end_time_dt = datetime.strptime(end_time, "%H:%M:%S")
						# Night Shift Handling
						# Check if the start time starts after 20:00
						if start_time_dt > datetime.strptime("20:00:00", "%H:%M:%S"):
							# Swap start time's A.M. and P.M.
							start_time_dt -= timedelta(hours=12)
						# Check if the end time is before 06:00
						if end_time_dt < datetime.strptime("06:00:00", "%H:%M:%S"):
							# Swap end time's A.M. and P.M.
							end_time_dt += timedelta(hours=12)
						# Calculate the duration between start_time and end_time
						duration = end_time_dt - start_time_dt
						# Check if the duration is greater than 4 hours
						if duration > timedelta(hours=4):
							# Add 45 minutes from start_time
							start_time_dt += timedelta(minutes=45)
							# Update start_time with the new value
							start_time = start_time_dt.strftime('%H:%M:%S')
					# Construct the formatted date
					current_year = datetime.now().year
					date_str = f"{current_year}-{sheet_name}"
					formatted_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%-m/%d/%Y")
					extracted_data.append([employee_no, "操作员", formatted_date, start_time, end_time, ""])

	# Load the import template workbook
	import_template_wb = load_workbook(import_template_file)
	import_template_ws = import_template_wb.active

	# Append the extracted data to the import template
	for row in extracted_data:
		import_template_ws.append(row)

	# Save the updated import template
	new_import_filename = f'importTemplate-{date}.xlsx'
	import_template_wb.save(os.path.join(output_dir, new_import_filename))

	st.success(f"Imported records saved to {os.path.join(output_dir, new_import_filename)}")

	# Provide a download button for the saved file
	with open(os.path.join(output_dir, new_import_filename), "rb") as file:
		btn = st.download_button(
			label="Download Generated Records",
			data=file,
			file_name=new_import_filename,
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
		)
	# Reset date
	date = None
